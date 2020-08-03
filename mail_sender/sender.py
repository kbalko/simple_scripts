#! python3
# -*- coding: utf-8 -*-
import openpyxl, smtplib, sys, os

# export var in CLI 
your_email = os.environ.get('YOUR_EMAIL')
password = os.environ.get('PASSWORD')

# read file
records = openpyxl.load_workbook('records.xlsx')
sheet = records.get_sheet_by_name('Sheet1')

last_col = sheet.max_column
last_month = sheet.cell(row=1, column=last_col).value

unpaid = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != u'zapłacono':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid[name] = email

#  email config
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login(your_email, password)  #  credencials

# send email
for name, email in unpaid.items():
    body = u"Subject: Zaległa wpłata za %s.\nCześć, %s!\nPrzypominam o zaległej wpłacie za %s" % (last_month, name, last_month)
    print(u"Wysłanie wiadomości e-mail na adres %s" % email)
    sendmail_status = smtp_obj.sendmail('drybrushlands@gmail.com', email, body.encode('utf-8'))
    if sendmail_status != {}:
        print(u"Wystąpił błąd podczas wysyłania wiadomości na adres %s:%s" % (email, sendmail_status))
smtp_obj.quit()
